import base64
import ctypes
from datetime import datetime
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, filedialog
from jinja2 import Template
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import pandas as pd
import requests
import winreg

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
:root { --verde: #008778; --verde-oscuro: #045c4f; --lima: #C4D70F; }
* { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; box-sizing: border-box; }
html, body { margin: 0 !important; padding: 0 !important; width: 100%; height: 100%; }
body { font-family: 'Inter', sans-serif; background-color: #f8fafc; color: #1e293b; font-size: 10px; }
@page { margin: 0px !important; size: auto; }
table { width: 100%; border-collapse: collapse; background: white; }
th { background: #ECF5E8; color: #008778; padding: 6px; text-align: left; border-bottom: 2px solid #C4D70F; font-size: 8px; text-transform: uppercase; }
td { border-bottom: 1px solid #ECF5E8; padding: 6px; font-size: 8px; }
tr { page-break-inside: avoid; }
.kpi-card { background: white; border-radius: 8px; padding: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); border-top: 4px solid var(--verde); }
.footer { font-size: 7px; color: #98999B; opacity: 0.7; text-transform: uppercase; margin-top: 20px; border-top: 1px solid #ddd; padding-top: 5px; display: flex; justify-content: space-between; }
</style>
</head>
<body>
<header style="background: linear-gradient(135deg, {{ color_hex }} 0%, {{ color_hex_end }} 100%); color: white; padding: 25px 30px; border-bottom: 5px solid #C4D70F; display: flex; justify-content: space-between; align-items: flex-start; margin: 0;">
    <div>
        <p style="font-size: 9px; letter-spacing: 2px; text-transform: uppercase; margin: 0; opacity: 0.8;">REPORTE DE CALIDAD</p>
        <h1 style="font-size: 24px; margin: 5px 0 0 0; text-transform: uppercase;">{{ titulo_reporte }}</h1>
        <p style="font-size: 10px; margin: 10px 0 0 0; color: #C4D70F; font-weight: bold; text-transform: uppercase;">GENERADO POR: {{ autor_nombre }}</p>
    </div>
    <div style="text-align: right;">
        <p style="font-size: 12px; font-weight: bold; margin: 0;">{{ periodo }}{% if subtitulo_turno %} | {{ subtitulo_turno }}{% endif %}</p>
        <p style="font-size: 9px; margin: 5px 0 0 0; opacity: 0.8;">FECHA DE EMISIÓN: {{ fecha_generacion }}</p>
    </div>
</header>
<div style="padding: 20px 30px;">
{% if modo == '2' %}
<div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin-bottom: 15px;">
    <div class="kpi-card" style="text-align: center;">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">Meta Producción</p>
        <p style="font-size: 24px; font-weight: bold; margin: 0; color: #008778;">{{ total_meta }}</p>
    </div>
    <div class="kpi-card" style="text-align: center;">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">Entregas Reales</p>
        <p style="font-size: 24px; font-weight: bold; margin: 0; color: #008778;">{{ total_entregado }}</p>
    </div>
    <div class="kpi-card" style="text-align: center; border-color: {{ color_borde_cump_hex }}; background-color: {{ bg_cump }};">
        <p style="font-size: 9px; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold; color: {{ color_borde_cump_hex }};">% CUMPLIMIENTO</p>
        <p style="font-size: 28px; font-weight: bold; margin: 0; color: {{ color_borde_cump_hex }};">{{ cumplimiento_str }}</p>
    </div>
</div>
{% endif %}
<div style="display: grid; grid-template-columns: repeat({{ kpi_cols }}, 1fr); gap: 15px; margin-bottom: 20px;">
    <div class="kpi-card" style="border-color: {{ color_hex }};">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">Volumen Global</p>
        <p style="font-size: 24px; font-weight: bold; margin: 0;">{{ total_unid }}</p>
    </div>
    {% if mostrar_rechazos %}
    <div class="kpi-card" style="border-color: #ef4444; background: #fef2f2;">
        <p style="font-size: 9px; color: #b91c1c; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">{% if modo == '2' %}Tasa Rechazo{% else %}Total Rechazos{% endif %}</p>
        <p style="font-size: 24px; font-weight: bold; color: #ef4444; margin: 0;">{% if modo == '2' %}{{ tasa_rechazo_str }}{% else %}{{ total_rechazos }}{% endif %}</p>
    </div>
    {% endif %}
    {% if mostrar_reparaciones %}
    <div class="kpi-card" style="border-color: #008778; background: #f0fdfa;">
        <p style="font-size: 9px; color: #0f766e; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">{% if modo == '2' %}Tasa Reparación{% else %}Total Reparaciones{% endif %}</p>
        <p style="font-size: 24px; font-weight: bold; color: #008778; margin: 0;">{% if modo == '2' %}{{ tasa_reparacion_str }}{% else %}{{ total_reparaciones }}{% endif %}</p>
    </div>
    {% endif %}
    <div class="kpi-card" style="border-color: {{ color_hex }};">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">{% if modo == '5' %}Familias Afectadas{% elif modo == '2' %}Incidencia Global{% else %}Tipos Defecto{% endif %}</p>
        <p style="font-size: 24px; font-weight: bold; margin: 0;">{% if modo == '2' %}{{ tasa_kpi_str }}{% else %}{{ total_tipos }}{% endif %}</p>
    </div>
</div>
{% if modo == '1' %}
<div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px;">
    {% if top_def_rech != 'N/A' and mostrar_rechazos %}
    <div class="kpi-card" style="border-color: #ef4444; padding: 10px;">
        <p style="font-size: 9px; color: #b91c1c; margin: 0 0 8px 0; text-transform: uppercase; font-weight: bold; border-bottom: 1px solid #fecaca; padding-bottom: 4px;">Dominancia Rechazo</p>
        <div style="display: flex; justify-content: space-between;">
            <div style="width: 50%; padding-right: 5px; border-right: 1px solid #fecaca;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">DEFECTO PRINCIPAL</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_def_rech }}</p>
            </div>
            <div style="width: 50%; padding-left: 5px;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">CAUSA RAÍZ</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_causa_rech }}</p>
            </div>
        </div>
    </div>
    {% endif %}
    {% if top_def_rep != 'N/A' and mostrar_reparaciones %}
    <div class="kpi-card" style="border-color: #008778; padding: 10px;">
        <p style="font-size: 9px; color: #0f766e; margin: 0 0 8px 0; text-transform: uppercase; font-weight: bold; border-bottom: 1px solid #ccfbf1; padding-bottom: 4px;">Dominancia Reparación</p>
        <div style="display: flex; justify-content: space-between;">
            <div style="width: 50%; padding-right: 5px; border-right: 1px solid #ccfbf1;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">DEFECTO PRINCIPAL</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_def_rep }}</p>
            </div>
            <div style="width: 50%; padding-left: 5px;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">CAUSA RAÍZ</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_causa_rep }}</p>
            </div>
        </div>
    </div>
    {% endif %}
    {% if top_def_mp != 'N/A' and mostrar_rechazos %}
    <div class="kpi-card" style="border-color: #d97706; padding: 10px;">
        <p style="font-size: 9px; color: #b45309; margin: 0 0 8px 0; text-transform: uppercase; font-weight: bold; border-bottom: 1px solid #fef3c7; padding-bottom: 4px;">Dominancia Mat. Prima</p>
        <div style="display: flex; justify-content: space-between;">
            <div style="width: 50%; padding-right: 5px; border-right: 1px solid #fef3c7;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">DEFECTO PRINCIPAL</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_def_mp }}</p>
            </div>
            <div style="width: 50%; padding-left: 5px;">
                <p style="font-size: 7px; color: #666; margin: 0 0 2px 0;">CAUSA RAÍZ</p>
                <p style="font-size: 10px; font-weight: bold; margin: 0; color: #333;">{{ top_causa_mp }}</p>
            </div>
        </div>
    </div>
    {% endif %}
</div>
{% endif %}
{% if modo == '3' or modo == '6' %}
<div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-bottom: 20px;">
    <div class="kpi-card" style="border-color: {{ color_hex }}; text-align: center;">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">{% if modo == '6' %}Volumen Selección{% else %}Categoría (M) Dominante{% endif %}</p>
        <p style="font-size: 18px; font-weight: bold; margin: 0; color: {{ color_hex }};">{% if modo == '6' %}{{ total_unid }}{% else %}{{ top_m }}{% endif %}</p>
    </div>
    <div class="kpi-card" style="border-color: {{ color_hex }}; text-align: center;">
        <p style="font-size: 9px; color: #666; margin: 0 0 5px 0; text-transform: uppercase; font-weight: bold;">{% if modo == '6' %}Defectos Comparados{% else %}Origen Específico N°1{% endif %}</p>
        <p style="font-size: 14px; font-weight: bold; margin: 0;">{% if modo == '6' %}{{ top_m }}{% else %}{{ top_causa }}{% endif %}</p>
    </div>
</div>
{% endif %}
<div style="display: flex; flex-direction: column; gap: 20px; margin-bottom: 20px;">
    {% for chart in graficos_dinamicos %}
    <div style="background: white; padding: 15px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
        <h3 style="font-size: 11px; margin-top: 0; color: {{ color_hex }}; border-bottom: 1px solid #eee; padding-bottom: 5px; text-transform: uppercase;">{{ chart.titulo }}</h3>
        <div style="text-align: center;"><img src="{{ chart.b64 }}" style="max-width: 100%; height: auto; max-height: 280px;"></div>
    </div>
    {% endfor %}
</div>
{% if usar_ia or (usar_obs and observaciones) %}
<div style="display: flex; flex-direction: column; gap: 15px; margin-bottom: 20px;">
    {% if usar_ia %}
    <div style="background: white; padding: 15px; border-left: 4px solid {{ color_hex }}; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
        <h3 style="font-size: 10px; margin-top: 0; color: {{ color_hex }}; text-transform: uppercase;">🤖 Análisis Asistido (Llama-3)</h3>
        <p style="font-size: 10px; line-height: 1.4; margin: 0; color: #333;">{{ analisis_ia | safe }}</p>
    </div>
    {% endif %}
    {% if usar_obs and observaciones %}
    <div style="background: white; padding: 15px; border-left: 4px solid #F5911E; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
        <h3 style="font-size: 10px; margin-top: 0; color: #F5911E; text-transform: uppercase;">📝 Notas</h3>
        <p style="font-size: 10px; line-height: 1.4; margin: 0; color: #333;">{{ observaciones | safe }}</p>
    </div>
    {% endif %}
</div>
{% endif %}
<div style="background: white; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow: hidden; margin-bottom: 20px;">
    <table>
        <thead>
            <tr>
                <th style="width: 10%;">TIPO</th>
                <th style="width: 10%;">FECHA</th>
                <th style="width: 10%;">ÁREA</th>
                <th style="width: 10%;">ORDEN</th>
                <th style="width: 10%;">PTA</th>
                <th style="width: 20%;">DEFECTO / FAMILIA</th>
                <th style="width: 25%;">CAUSA</th>
                <th style="width: 5%; text-align: center;">CANT</th>
            </tr>
        </thead>
        <tbody>
            {% for f in datos_tabla %}
            <tr>
                <td style="font-weight: bold; {% if 'RECHAZO MP' in f.TIPO %}color: #d97706;{% elif 'RECHAZO' in f.TIPO %}color: #ef4444;{% else %}color: #008778;{% endif %}">{{ f.TIPO }}</td>
                <td>{{ f.FECHA }}</td>
                <td>{{ f.AREA }}</td>
                <td style="color: {{ color_hex }}; font-weight: bold;">{{ f.ORDEN }}</td>
                <td>{{ f.PTA }}</td>
                <td style="font-weight: bold;">{{ f.DEFECTO }}</td>
                <td>{{ f.CAUSA }}</td>
                <td style="text-align: center; font-weight: bold; font-size: 9px; background: #f8fafc;">{{ f.CANT }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
<div class="footer">
    <span>Maforsa</span>
    <span>Gestión de reportes</span>
</div>
</div>
</body>
</html>
"""

def usar_modo_oscuro():
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'SOFTWARE\Microsoft\Windows\CurrentVersion\Themes\Personalize') as key:
            valor = winreg.QueryValueEx(key, 'AppsUseLightTheme')[0]
            return valor == 0
    except Exception:
        return False

def ruta_recurso(rel_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, rel_path)

def aplicar_tema_ventana(ventana):
    try:
        ventana.update_idletasks()
        hwnd = ctypes.windll.user32.GetParent(ventana.winfo_id())
        if not hwnd:
            hwnd = ventana.winfo_id()
        valor = ctypes.c_int(1 if usar_modo_oscuro() else 0)
        for attr in (20, 19):
            if ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, attr, ctypes.byref(valor), ctypes.sizeof(valor)) == 0:
                break
    except Exception:
        pass
    try:
        ruta_icono = ruta_recurso("icon.ico")
        if os.path.exists(ruta_icono):
            ventana.iconbitmap(default=ruta_icono)
        elif sys.executable.endswith('.exe'):
            ventana.iconbitmap(default=sys.executable)
    except Exception:
        pass

def obtener_ruta_escritorio():
    user_home = os.path.expanduser("~")
    posibles_rutas = [
        os.path.join(user_home, "OneDrive", "Desktop"),
        os.path.join(user_home, "OneDrive", "Escritorio"),
        os.path.join(user_home, "Desktop"),
        os.path.join(user_home, "Escritorio")
    ]
    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            return ruta
    return user_home

def obtener_ruta_config_autor():
    return os.path.join(os.path.expanduser("~"), ".mafor_autor.txt")

RUTA_LOG = os.path.join(obtener_ruta_escritorio(), "Mafor_Reportes_Log.txt")
RUTA_EXCEL_LOCAL = os.path.join(obtener_ruta_escritorio(), "Mafor_Datos_Sync.xlsx")

def registrar_error(contexto, ex):
    try:
        with open(RUTA_LOG, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%d-%m-%Y %H:%M:%S')}] {contexto}: {ex}\n")
    except Exception:
        pass

def normalizar_fecha(v):
    if pd.isna(v):
        return pd.NaT
    if isinstance(v, (pd.Timestamp, datetime)):
        return pd.to_datetime(v).normalize()
    s = str(v).strip().split(' ')[0]
    try:
        return pd.to_datetime(s, dayfirst=True).normalize()
    except Exception:
        s = s.lower()
        meses = {
            'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04',
            'may': '05', 'jun': '06', 'jul': '07', 'ago': '08',
            'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'
        }
        for t, n in meses.items():
            if t in s:
                s = s.replace(t, n)
                break
        if len(s) <= 6:
            s = s.strip('-/') + f"-{datetime.now().year}"
        return pd.to_datetime(s, errors='coerce', dayfirst=True).normalize()

def clasificar_6m(causa):
    map_6m = {
        'MANO DE OBRA': ['manipulacion', 'armado', 'operador', 'humano', 'golpe', 'mal', 'descuido', 'caida'],
        'MÁQUINA': ['falla maquina', 'fresadora', 'cnc', 'prensa', 'escuadra', 'rodillo'],
        'MATERIAL': ['nudo', 'humedad', 'terciado', 'pino', 'mdf', 'mp', 'materia prima', 'despegue', 'folio', 'adhesivo'],
        'MÉTODO': ['procedimiento', 'estandar', 'apilamiento', 'ruta'],
        'MEDIO AMBIENTE': ['temperatura', 'polvo', 'suciedad'],
        'MEDIDA': ['dimension', 'tolerancia', 'fuera de medida', 'escuadria']
    }
    c = str(causa).lower()
    for k, v_list in map_6m.items():
        if any(v in c for v in v_list):
            return k
    return 'MÉTODO'

def mapear_tipo_incidencia(val):
    if pd.isna(val):
        return 'RECHAZO'
    s = str(val).strip().upper()
    if 'REPAR' in s or 'REP' in s:
        return 'REPARACION'
    if 'RECH' in s or 'NO CONF' in s:
        return 'RECHAZO'
    return 'RECHAZO'

def limpiar_dataframe(df, tipo_incidencia=None):
    def find_col(keywords, df_cols, exclude=None):
        for col in df_cols:
            if exclude and any(ex in col for ex in exclude):
                continue
            for kw in keywords:
                if kw in col:
                    return col
        return None

    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [
            dup + '_' + str(i) if i != 0 else dup for i in range(sum(cols == dup))
        ]
    df.columns = cols
    df.columns = df.columns.astype(str).str.strip().str.upper().str.replace('\n', ' ')
    
    tipo_orig_col = None
    col_estatus = find_col(['ESTAT', 'ESTADO', 'DISPOSICION', 'RESULTADO', 'TIPO'], df.columns, exclude=['INCIDENCIA', 'HALLAZGO'])
    
    if col_estatus:
        tipo_orig_col = df[col_estatus].copy()
        df = df.drop(columns=[col_estatus])
        
    if 'TIPO_INCIDENCIA' in df.columns:
        df = df.drop(columns=['TIPO_INCIDENCIA'])
    if 'TIPO' in df.columns:
        df = df.drop(columns=['TIPO'])
        
    renombrado_hecho = False
    for col in list(df.columns):
        if "FECHA" in col and not renombrado_hecho:
            df.rename(columns={col: "FECHA REVISION"}, inplace=True)
            renombrado_hecho = True
            
    col_fecha = find_col(['FECHA', 'ECHA REVIS'], df.columns, exclude=['SOL', 'ING'])
    if col_fecha:
        df.rename(columns={col_fecha: 'FECHA REVISION'}, inplace=True)
        
    col_area = find_col(['ARE'], df.columns, exclude=['SIG'])
    if col_area:
        df.rename(columns={col_area: 'AREA'}, inplace=True)
        
    col_defecto = find_col(['DEFECT'], df.columns, exclude=['CODIGO', 'CÓDIGO'])
    if not col_defecto:
        col_defecto = find_col(['HALLAZGO'], df.columns)
    if col_defecto:
        df.rename(columns={col_defecto: 'DEFECTO'}, inplace=True)
        
    col_causa = find_col(['CAUSA'], df.columns)
    if not col_causa:
        col_causa = find_col(['OBSERVACION'], df.columns)
    if col_causa:
        df.rename(columns={col_causa: 'POSIBLE CAUSA RAIZ'}, inplace=True)
        
    col_cant = find_col(['CANT'], df.columns)
    if col_cant:
        df.rename(columns={col_cant: 'CANTIDAD'}, inplace=True)
        
    col_fam = find_col(['FAMIL'], df.columns)
    if col_fam:
        df.rename(columns={col_fam: 'FAMILIA'}, inplace=True)
        
    col_neg = find_col(['NEGOC', 'RETAIL'], df.columns)
    if col_neg:
        df.rename(columns={col_neg: 'NEGOCIO'}, inplace=True)
        
    col_resp = find_col(['RESPONSABLE'], df.columns)
    if col_resp:
        df.rename(columns={col_resp: 'RESPONSABLE TURNO'}, inplace=True)
        
    col_turno = find_col(['TURN'], df.columns, exclude=['CLAF', 'RESPONSABLE'])
    if col_turno:
        df.rename(columns={col_turno: 'TURNO'}, inplace=True)
        
    col_orden = find_col(['ORDEN', 'NV', 'LOTE'], df.columns)
    if col_orden:
        df.rename(columns={col_orden: 'ORDEN'}, inplace=True)
        
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index.values.tolist()] = [
            dup + '_' + str(i) if i != 0 else dup for i in range(sum(cols == dup))
        ]
    df.columns = cols
    
    if 'TURNO' in df.columns:
        df['TURNO'] = df['TURNO'].astype(str).str.strip().str.upper()
        df['TURNO'] = df['TURNO'].map({'DIA': 'Turno A', 'NOCHE': 'Turno B'}).fillna(df['TURNO'])
        df['TURNO'] = df['TURNO'].apply(lambda x: 'Turno A' if x == 'A' else ('Turno B' if x == 'B' else x))
        
    if 'AREA' in df.columns:
        mapa_areas = {
            '1': 'ARMADO', '2': 'ESCUADRA', '3': 'DIMENSIONADO/CNC', '4': 'RANURADO',
            '5': 'PINTURA', '6': 'PRECOLGADO', '7': 'POMO', '8': 'CARPINTERIA',
            '9': 'EMBOLSADO', '10': 'FOLIADORA', '11': 'BODEGA', '12': 'DESPACHO/EMBALAJE',
            '13': 'EXPORTACION', '14': 'REPARACIONES', '15': 'TROZADO', '16': 'PRENSA',
            '17': 'TRANSPORTISTA'
        }
        df['AREA'] = df['AREA'].fillna('').astype(str).str.strip().apply(lambda x: mapa_areas.get(x.split('.')[0], x.upper()))
        
    if 'FECHA REVISION' in df.columns:
        df['FECHA REVISION'] = df['FECHA REVISION'].apply(normalizar_fecha)
        
    if tipo_incidencia is not None:
        df['TIPO_INCIDENCIA'] = tipo_incidencia
    elif tipo_orig_col is not None:
        df['TIPO_INCIDENCIA'] = tipo_orig_col.apply(mapear_tipo_incidencia)
    else:
        df['TIPO_INCIDENCIA'] = 'RECHAZO'
        
    if 'DEFECTO' in df.columns and tipo_incidencia != 'REPA_PROD':
        df.dropna(subset=['DEFECTO'], inplace=True)
        df = df[df['DEFECTO'].astype(str).str.strip() != '']
        df = df[df['DEFECTO'].astype(str).str.strip().str.upper() != 'NAN']
        
    if 'CANTIDAD' in df.columns:
        df['CANTIDAD'] = pd.to_numeric(df['CANTIDAD'], errors='coerce').fillna(0)
        df = df[df['CANTIDAD'] > 0]
        
    return df

def exportar_excel_calidad(df_infografia, df_pnc_completo, ruta_directorio, nombre_base):
    ruta_excel = os.path.join(ruta_directorio, f"{nombre_base}_Datos.xlsx")
    df_info = df_infografia.copy()
    
    col_responsable = 'RESPONSABLE TURNO' if 'RESPONSABLE TURNO' in df_info.columns else ('RESPONSABLE' if 'RESPONSABLE' in df_info.columns else 'TURNO')
    col_defecto = 'DEFECTO' if 'DEFECTO' in df_info.columns else ('HALLAZGO' if 'HALLAZGO' in df_info.columns else (df_info.columns[0] if not df_info.empty else 'DEFECTO'))
    col_estatus = 'TIPO_INCIDENCIA' if 'TIPO_INCIDENCIA' in df_info.columns else ('ESTATUS' if 'ESTATUS' in df_info.columns else (df_info.columns[1] if not df_info.empty else 'ESTATUS'))
    
    df_pnc = df_pnc_completo.copy()
    if 'FECHA REVISION' in df_info.columns:
        df_info['FECHA REVISION'] = pd.to_datetime(df_info['FECHA REVISION'], errors='coerce').dt.strftime('%d-%m-%Y')
    if not df_pnc.empty and 'FECHA REVISION' in df_pnc.columns:
        df_pnc['FECHA REVISION'] = pd.to_datetime(df_pnc['FECHA REVISION'], errors='coerce').dt.strftime('%d-%m-%Y')
        
    try:
        tabla_dinamica = pd.pivot_table(df_info, values='CANTIDAD', index=[col_responsable, col_defecto], columns=[col_estatus], aggfunc='sum', fill_value=0)
        if 'REPARACION' in tabla_dinamica.columns and 'REPARABLE' not in tabla_dinamica.columns:
            tabla_dinamica = tabla_dinamica.rename(columns={'REPARACION': 'REPARABLE'})
        for c_status in ['RECHAZO', 'REPARABLE']:
            if c_status not in tabla_dinamica.columns:
                tabla_dinamica[c_status] = 0
        tabla_dinamica['TOTAL'] = tabla_dinamica.sum(axis=1)
    except Exception:
        try:
            tabla_dinamica = df_info.groupby([col_responsable, col_defecto])['CANTIDAD'].sum().to_frame(name='TOTAL')
        except Exception:
            tabla_dinamica = pd.DataFrame()
            
    try:
        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            if not df_info.empty:
                df_info.to_excel(writer, sheet_name='Datos_Infografia', index=False)
            else:
                pd.DataFrame([["Sin Datos"]]).to_excel(writer, sheet_name='Datos_Infografia', index=False, header=False)
                
            if not tabla_dinamica.empty:
                tabla_dinamica.to_excel(writer, sheet_name='Resumen_Dinamico')
            else:
                pd.DataFrame([["Sin Datos"]]).to_excel(writer, sheet_name='Resumen_Dinamico', index=False, header=False)
                
            if not df_pnc.empty:
                df_pnc.to_excel(writer, sheet_name='Historico_PNC', index=False)
            else:
                pd.DataFrame([["Sin Datos"]]).to_excel(writer, sheet_name='Historico_PNC', index=False, header=False)
                
            workbook = writer.book
            borde_tenue = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
            fondo_header = PatternFill(start_color='ECF5E8', end_color='ECF5E8', fill_type='solid')
            fuente_header = Font(bold=True, color='008778')
            
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                for col_cells in ws.columns:
                    col_letter = col_cells[0].column_letter
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            cell.border = borde_tenue
                            cell.alignment = Alignment(vertical='center', wrap_text=False)
                            val_str = str(cell.value)
                            if len(val_str) > max_len:
                                max_len = len(val_str)
                            if sheet_name == 'Resumen_Dinamico':
                                if cell.value in ['TOTAL', 'RECHAZO', 'REPARABLE', col_responsable, col_defecto]:
                                    cell.fill = fondo_header
                                    cell.font = fuente_header
                            else:
                                if isinstance(cell.value, str) and ((sheet_name == 'Datos_Infografia' and cell.value in df_info.columns) or (sheet_name == 'Historico_PNC' and cell.value in df_pnc.columns)):
                                    cell.fill = fondo_header
                                    cell.font = fuente_header
                ws.column_dimensions[col_letter].width = min(max_len + 8, 50)
    except Exception:
        pass

def convertir_html_a_pdf(ruta_html, ruta_pdf):
    rutas_edge = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        "msedge"
    ]
    ejecutable = next((p for p in rutas_edge if os.path.exists(p) or p == "msedge"), "msedge")
    user_data_dir = tempfile.mkdtemp(prefix="edge_pdf_")
    ruta_html_abs = os.path.abspath(ruta_html).replace("\\", "/")
    url_html = f"file:///{ruta_html_abs}"
    comando = [
        ejecutable, "--headless", "--disable-gpu", "--no-pdf-header-footer",
        "--print-to-pdf-no-margin", f"--user-data-dir={user_data_dir}",
        f"--print-to-pdf={os.path.abspath(ruta_pdf)}", url_html
    ]
    try:
        subprocess.run(comando, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True)
        time.sleep(1)
    finally:
        if os.path.exists(user_data_dir):
            shutil.rmtree(user_data_dir, ignore_errors=True)

def aplicar_metadata_pdf(ruta_pdf, autor):
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(ruta_pdf)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        writer.add_metadata({"/Author": autor, "/Creator": "Sistema Mafor - Gestión de Calidad"})
        with open(ruta_pdf, "wb") as f:
            writer.write(f)
    except Exception:
        pass

def redactar_analisis_ia(datos_contexto, api_key):
    if not api_key: return "<i style='color:#ef4444; font-weight:bold;'>[Falta API Key de Groq]</i>"
    system_prompt = (
        "Eres un analista de calidad senior en Mafor SA. Redactas resúmenes ejecutivos "
        "breves para gerencia a partir de datos de indicadores de calidad.\n\n"
        "Reglas de estilo (aplican siempre):\n"
        "- Tono corporativo, directo y objetivo.\n"
        "- Nunca uses saludos, introducciones ni cierres (ej: 'Aquí tienes...', 'En resumen...').\n"
        "- Responde solo con texto plano: sin HTML, sin markdown, sin viñetas ni títulos.\n"
        "- Separa los 2 párrafos con una línea en blanco.\n"
        "- Responde siempre en español, sin mezclar palabras en inglés.\n"
        "- Extensión total: no más de ~70 palabras entre los 2 párrafos. Prioriza claridad "
        "sobre cumplir un número exacto de palabras."
    )
    user_prompt = (
        f"Datos de calidad del período:\n{datos_contexto}\n\n"
        "Cubre, en este orden:\n"
        "1. Si la tasa de incidencias está dentro o fuera del rango esperado frente a la meta "
        "(usa 'meta_tasa_incidencias_pct' si viene en los datos; si no viene, no afirmes "
        "dentro/fuera de rango, solo reporta la tasa actual).\n"
        "2. Si el problema está concentrado en pocos defectos/causas o disperso entre varios.\n"
        "3. UNA recomendación concreta y accionable, ligada directamente al defecto o causa principal.\n\n"
        "Si algún dato no permite evaluar un punto, dilo brevemente en vez de inventar información."
    )
    intentos=0
    while intentos<3:
        try:
            url_api="https://api.groq.com/openai/v1/chat/completions"
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
            response=requests.post(url_api, headers=headers, json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                "temperature": 0.3,
            })
            if response.status_code==200: return response.json()['choices'][0]['message']['content'].strip().replace('\n', '<br>')
            else:
                intentos+=1
                time.sleep(3)
        except Exception:
            intentos+=1
            time.sleep(3)
    return "<i style='color:#ef4444; font-weight:bold;'>[Error Técnico IA]</i>"

def get_b64(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', transparent=True, dpi=120)
    plt.close(fig)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode('utf-8')

def plot_bar(labels, datasets, stacked=False, figsize=(8, 3.5)):
    fig, ax = plt.subplots(figsize=figsize)
    x = np.arange(len(labels))
    if stacked:
        bottom = np.zeros(len(labels))
        totals = np.zeros(len(labels))
        for ds in datasets:
            totals += np.array(ds['data'])
        global_max = max(totals) if len(totals) > 0 else 1
        threshold = global_max * 0.06 
        for ds in datasets:
            data_arr = np.array(ds['data'])
            if sum(data_arr) > 0:
                bars = ax.bar(x, data_arr, label=ds['label'], bottom=bottom, color=ds['color'])
                for bar in bars:
                    h = bar.get_height()
                    if h > threshold:
                        ax.text(bar.get_x() + bar.get_width() / 2., bar.get_y() + h / 2., f'{int(h)}', ha='center', va='center', color='white', fontsize=7, fontweight='bold')
                bottom += data_arr
        ax.set_ylim(0, global_max * 1.15)
    else:
        valid_ds = [ds for ds in datasets if sum(ds['data']) > 0]
        n = len(valid_ds)
        if n > 0:
            width = 0.8 / n
            max_h = max([max(ds['data']) for ds in valid_ds]) if valid_ds else 1
            ax.set_ylim(0, max_h * 1.25)
            for i, ds in enumerate(valid_ds):
                pos = x + i * width - width / 2 + width / (2 * n)
                bars = ax.bar(pos, ds['data'], width, label=ds['label'], color=ds['color'])
                for bar in bars:
                    h = bar.get_height()
                    if h > 0:
                        ax.text(bar.get_x() + bar.get_width() / 2., h + (max_h * 0.02), f'{int(h)}', ha='center', va='bottom', color='black', fontsize=7)
    ax.set_xticks(x)
    short_labels = [str(lbl)[:15] + '..' if len(str(lbl)) > 15 else str(lbl) for lbl in labels]
    ax.set_xticklabels(short_labels, rotation=45, ha='right', fontsize=8)
    if len(datasets) > 1:
        ax.legend(fontsize=8, loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=3, frameon=False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.3)
    fig.tight_layout()
    return get_b64(fig)

def plot_line(labels, datasets, figsize=(8, 3.5)):
    fig, ax = plt.subplots(figsize=figsize)
    global_max = max([max(ds['data']) for ds in datasets if sum(ds['data']) > 0]) if datasets else 1
    offset = global_max * 0.05
    step = max(1, len(labels) // 15)
    for i in range(len(labels)):
        points_at_x = []
        for ds in datasets:
            if sum(ds['data']) > 0 and ds['data'][i] > 0:
                if step <= 2 or i % step == 0 or ds['data'][i] == max(ds['data']):
                    points_at_x.append((ds['data'][i], ds['color']))
        points_at_x.sort(key=lambda x: x[0], reverse=True)
        occupied_y = []
        for y_val, color in points_at_x:
            label_y = y_val + offset
            while any(abs(label_y - prev_y) < (global_max * 0.08) for prev_y in occupied_y):
                label_y -= (global_max * 0.08)
            occupied_y.append(label_y)
            ax.text(i, label_y, f'{int(y_val)}', ha='center', va='center', fontsize=7, color=color, fontweight='bold')
    for ds in datasets:
        if sum(ds['data']) > 0:
            ax.plot(labels, ds['data'], label=ds['label'], color=ds['color'], marker='o', lw=2)
    ax.set_xticks(range(0, len(labels), step))
    ax.set_xticklabels([labels[i] for i in range(0, len(labels), step)], rotation=45, ha='right', fontsize=8)
    if len(datasets) > 1:
        ax.legend(fontsize=8, loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=3, frameon=False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.3)
    y_min, y_max = ax.get_ylim()
    ax.set_ylim(y_min, max(y_max, global_max * 1.2))
    fig.tight_layout()
    return get_b64(fig)

def plot_pie(labels, sizes, figsize=(5, 5)):
    fig, ax = plt.subplots(figsize=figsize)
    colors = ['#008778', '#ef4444', '#F5911E', '#3b82f6', '#8b5cf6', '#14b8a6']
    ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors, textprops={'fontsize': 8})
    ax.axis('equal')
    fig.tight_layout()
    return get_b64(fig)

def plot_pareto(labels, data, figsize=(8, 3.5)):
    fig, ax1 = plt.subplots(figsize=figsize)
    sorted_idx = np.argsort(data)[::-1]
    data_sorted = np.array(data)[sorted_idx]
    labels_sorted = np.array(labels)[sorted_idx]
    bars = ax1.bar(range(len(data_sorted)), data_sorted, color='#008778')
    ax1.set_xticks(range(len(labels_sorted)))
    short_labels = [str(lbl)[:15] + '..' if len(str(lbl)) > 15 else str(lbl) for lbl in labels_sorted]
    ax1.set_xticklabels(short_labels, rotation=45, ha='right', fontsize=8)
    cum_sum = np.cumsum(data_sorted)
    if cum_sum[-1] > 0:
        cum_perc = cum_sum / cum_sum[-1] * 100
    else:
        cum_perc = np.zeros(len(data_sorted))
    ax2 = ax1.twinx()
    ax2.plot(range(len(data_sorted)), cum_perc, color='#ef4444', marker='o', ms=4, lw=2)
    ax2.set_ylim(0, 110)
    ax1.spines['top'].set_visible(False)
    ax2.spines['top'].set_visible(False)
    for i, p in enumerate(cum_perc):
        ax2.text(i, p + 3, f'{int(p)}%', ha='center', va='bottom', fontsize=7, color='#ef4444')
    fig.tight_layout()
    return get_b64(fig)

def plot_heatmap(df, col_y, col_x, figsize=(8, 4)):
    fig, ax = plt.subplots(figsize=figsize)
    pt = pd.pivot_table(df, values='CANTIDAD_FISICA', index=col_y, columns=col_x, aggfunc='sum', fill_value=0)
    if pt.empty:
        return ""
    ax.imshow(pt.values, cmap='Reds', aspect='auto')
    ax.set_xticks(np.arange(len(pt.columns)))
    ax.set_yticks(np.arange(len(pt.index)))
    short_cols = [str(c)[:10] + '..' if len(str(c)) > 10 else str(c) for c in pt.columns]
    short_idx = [str(i)[:15] + '..' if len(str(i)) > 15 else str(i) for i in pt.index]
    ax.set_xticklabels(short_cols, rotation=45, ha='right', fontsize=8)
    ax.set_yticklabels(short_idx, fontsize=8)
    for i in range(len(pt.index)):
        for j in range(len(pt.columns)):
            val = pt.values[i, j]
            if val > 0:
                color = 'white' if val > pt.values.max() * 0.5 else 'black'
                ax.text(j, i, int(val), ha='center', va='center', color=color, fontsize=7)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    fig.tight_layout()
    return get_b64(fig)

def plot_stacked_area(labels, datasets, figsize=(8, 3.5)):
    fig, ax = plt.subplots(figsize=figsize)
    x = range(len(labels))
    y_data = [ds['data'] for ds in datasets]
    colors = [ds['color'] for ds in datasets]
    lbls = [ds['label'] for ds in datasets]
    step = max(1, len(labels) // 15)
    if y_data and any(sum(d) > 0 for d in y_data):
        ax.stackplot(x, y_data, labels=lbls, colors=colors, alpha=0.8)
        ax.set_xticks(range(0, len(labels), step))
        ax.set_xticklabels([labels[i] for i in range(0, len(labels), step)], rotation=45, ha='right', fontsize=8)
        if len(datasets) > 1:
            ax.legend(fontsize=8, loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=3, frameon=False)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(axis='y', linestyle='--', alpha=0.3)
    fig.tight_layout()
    return get_b64(fig)

def plot_hbar(labels, data, color='#008778', figsize=(8, 4)):
    fig, ax = plt.subplots(figsize=figsize)
    if not labels or sum(data) <= 0:
        return ""
    order = np.argsort(data)
    labels_o = [str(labels[i])[:25] for i in order]
    data_o = [data[i] for i in order]
    y = np.arange(len(labels_o))
    bars = ax.barh(y, data_o, color=color)
    for bar in bars:
        w = bar.get_width()
        if w > 0:
            ax.text(w + max(data_o) * 0.01, bar.get_y() + bar.get_height() / 2, f'{int(w)}', va='center', fontsize=8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels_o, fontsize=8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.grid(axis='x', linestyle='--', alpha=0.3)
    fig.tight_layout()
    return get_b64(fig)

def plot_control(labels, data, figsize=(8, 3.5)):
    fig, ax = plt.subplots(figsize=figsize)
    if not labels or len(data) == 0:
        return ""
    arr = np.array(data, dtype=float)
    media = arr.mean()
    sigma = arr.std()
    lcs = media + 2 * sigma
    lci = max(media - 2 * sigma, 0)
    x = np.arange(len(labels))
    ax.plot(x, arr, marker='o', markersize=4, color='#008778', linewidth=1.5, label='Cantidad diaria', zorder=3)
    ax.axhline(media, color='#3b82f6', linestyle='--', linewidth=1, label=f'Promedio ({media:.1f})')
    ax.axhline(lcs, color='#ef4444', linestyle=':', linewidth=1, label=f'Límite superior ({lcs:.1f})')
    ax.axhline(lci, color='#ef4444', linestyle=':', linewidth=1)
    ax.fill_between(x, lci, lcs, color='#ef4444', alpha=0.06)
    fuera = arr > lcs
    if fuera.any():
        ax.scatter(x[fuera], arr[fuera], color='#ef4444', s=45, zorder=4, label='Fuera de control')
    step = max(1, len(labels) // 15)
    ax.set_xticks(range(0, len(labels), step))
    ax.set_xticklabels([labels[i] for i in range(0, len(labels), step)], rotation=45, ha='right', fontsize=8)
    ax.legend(fontsize=7, loc='upper center', bbox_to_anchor=(0.5, 1.18), ncol=4, frameon=False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.3)
    fig.tight_layout()
    return get_b64(fig)

class PanelControlMafor:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Información de Calidad - MAFOR SA")
        self.es_oscuro = usar_modo_oscuro()
        aplicar_tema_ventana(self.root)
        self.root.bind_all("<Map>", self._on_map_tema)
        self.root.geometry("1000x650")
        self.root.minsize(800, 600)
        
        bg_main = "#121212" if self.es_oscuro else "#f8fafc" 
        fg_main = "#ffffff" if self.es_oscuro else "#1e293b"
        accent = "#0d5c46" if self.es_oscuro else "#008778"
        accent_hover = "#0a4736" if self.es_oscuro else "#045c4f"
        entry_bg = "#242424" if self.es_oscuro else "#ffffff"
        entry_fg = "#ffffff" if self.es_oscuro else "#000000"
        
        self.root.configure(bg=bg_main, bd=0, highlightthickness=0)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        
        self.root.option_add('*TCombobox*Listbox.background', entry_bg)
        self.root.option_add('*TCombobox*Listbox.foreground', entry_fg)
        self.root.option_add('*TCombobox*Listbox.selectBackground', accent)
        self.root.option_add('*TCombobox*Listbox.selectForeground', 'white')
        
        self.defectos_seleccionados = []
        self.graficos_seleccionados = ["Gráfico Principal (Según Modo)", "Gráfico Evolutivo (Tiempo)"]
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background=bg_main, borderwidth=0)
        style.configure("TLabel", background=bg_main, foreground=fg_main, font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10, "bold"), background=accent, foreground="white", borderwidth=0, focuscolor=accent)
        style.map("TButton", background=[("active", accent_hover)])
        style.configure("TCheckbutton", background=bg_main, foreground=fg_main, font=("Segoe UI", 10), indicatorbackground=entry_bg, indicatorcolor=entry_bg)
        style.map("TCheckbutton", background=[("active", bg_main)], indicatorcolor=[("selected", accent)])
        style.configure("TCombobox", fieldbackground=entry_bg, background=entry_bg, foreground=entry_fg, borderwidth=0, arrowcolor=fg_main)
        style.map("TCombobox", fieldbackground=[("readonly", entry_bg)], selectbackground=[("readonly", accent)], selectforeground=[("readonly", "white")])
        style.configure("TEntry", fieldbackground=entry_bg, foreground=entry_fg, borderwidth=0)
        style.configure("Vertical.TScrollbar", gripcount=0, background=entry_bg, darkcolor=bg_main, lightcolor=bg_main, troughcolor=bg_main, bordercolor=bg_main, arrowcolor=fg_main)
        
        header_frame = tk.Frame(root, bg=accent, height=60, bd=0, highlightthickness=0)
        header_frame.grid(row=0, column=0, sticky="ew")
        header_frame.columnconfigure(0, weight=1)
        
        lbl_title = tk.Label(header_frame, text="SISTEMA DE REPORTES Y KPI DE CALIDAD", font=("Segoe UI", 14, "bold"), fg="white", bg=accent, bd=0, highlightthickness=0)
        lbl_title.pack(side="left", padx=20, pady=15)
        
        self.btn_sync = ttk.Button(header_frame, text="🔄 Descargar SharePoint", command=self.hilo_sincronizar)
        self.btn_sync.pack(side="right", padx=10, pady=15)
        
        self.btn_local = ttk.Button(header_frame, text="📂 Cargar Excel Local", command=self.cargar_excel_local)
        self.btn_local.pack(side="right", padx=10, pady=15)
        
        main_frame = tk.Frame(root, bg=bg_main, padx=20, pady=20, bd=0, highlightthickness=0)
        main_frame.grid(row=1, column=0, sticky="nsew")
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(13, weight=1)
        
        grid_lbl = {"padx": 10, "pady": 5, "sticky": "w"}
        grid_inp = {"padx": 10, "pady": 5, "sticky": "ew"}
        
        ttk.Label(main_frame, text="Generado por:").grid(row=0, column=0, **grid_lbl)
        self.txt_autor = ttk.Entry(main_frame)
        self.txt_autor.grid(row=0, column=1, **grid_inp)
        
        ruta_autor = obtener_ruta_config_autor()
        if os.path.exists(ruta_autor):
            try:
                with open(ruta_autor, 'r', encoding='utf-8') as f:
                    ultimo_autor = f.read().strip()
                self.txt_autor.insert(0, ultimo_autor)
            except Exception:
                pass
            
        ttk.Label(main_frame, text="Modo de Reporte:").grid(row=1, column=0, **grid_lbl)
        self.cb_modo = ttk.Combobox(main_frame, state="readonly", values=["1 - ESTÁNDAR", "2 - KPI CONSOLIDADO", "3 - ANÁLISIS 6M", "4 - EVOLUCIÓN DEFECTO N°1", "5 - ENFOQUE FAMILIA", "6 - COMPARATIVA EVOLUCIÓN DEFECTOS"])
        self.cb_modo.set("1 - ESTÁNDAR")
        self.cb_modo.grid(row=1, column=1, **grid_inp)
        
        ttk.Label(main_frame, text="Categoría de Análisis:").grid(row=2, column=0, **grid_lbl)
        self.cb_tipo = ttk.Combobox(main_frame, state="readonly", values=["1 - RECHAZOS", "2 - REPARACIONES", "3 - RECHAZOS MATERIA PRIMA", "4 - RECHAZO + REPARACION", "5 - CONSOLIDADO TOTAL"])
        self.cb_tipo.set("5 - CONSOLIDADO TOTAL")
        self.cb_tipo.grid(row=2, column=1, **grid_inp)
        
        ttk.Label(main_frame, text="Filtro de Turno:").grid(row=3, column=0, **grid_lbl)
        self.cb_turno = ttk.Combobox(main_frame, state="readonly", values=["1 - Ambos Turnos (Consolidado 24h)", "2 - Solo Turno A (Día)", "3 - Solo Turno B (Noche)"])
        self.cb_turno.set("1 - Ambos Turnos (Consolidado 24h)")
        self.cb_turno.grid(row=3, column=1, **grid_inp)
        
        ttk.Label(main_frame, text="Fecha Inicio (DD-MM-YYYY):").grid(row=4, column=0, **grid_lbl)
        self.txt_f_in = ttk.Entry(main_frame)
        self.txt_f_in.grid(row=4, column=1, **grid_inp)
        
        ttk.Label(main_frame, text="Fecha Fin (DD-MM-YYYY):").grid(row=5, column=0, **grid_lbl)
        self.txt_f_out = ttk.Entry(main_frame)
        self.txt_f_out.grid(row=5, column=1, **grid_inp)
        
        self.lbl_modo6_text = ttk.Label(main_frame, text="Modo 6:")
        self.frame_list = tk.Frame(main_frame, bg=bg_main, bd=0, highlightthickness=0)
        self.btn_cargar = ttk.Button(self.frame_list, text="Obtener y Seleccionar", command=self.hilo_cargar_defectos)
        self.btn_cargar.pack(side="left", pady=2)
        self.lbl_seleccionados = ttk.Label(self.frame_list, text="0 seleccionados")
        self.lbl_seleccionados.pack(side="left", padx=10)
        
        ttk.Label(main_frame, text="Formato de Salida:").grid(row=7, column=0, **grid_lbl)
        self.cb_formato = ttk.Combobox(main_frame, state="readonly", values=["PDF", "HTML", "Ambos (PDF + HTML)"])
        self.cb_formato.set("PDF")
        self.cb_formato.grid(row=7, column=1, **grid_inp)
        
        ttk.Label(main_frame, text="Gráficos a Incluir:").grid(row=8, column=0, **grid_lbl)
        frame_graf = tk.Frame(main_frame, bg=bg_main, bd=0, highlightthickness=0)
        frame_graf.grid(row=8, column=1, **grid_inp)
        self.btn_graficos = ttk.Button(frame_graf, text="📊 Seleccionar Gráficos", command=self.abrir_popup_graficos)
        self.btn_graficos.pack(side="left", pady=2)
        self.lbl_graficos = ttk.Label(frame_graf, text=f"{len(self.graficos_seleccionados)} seleccionados")
        self.lbl_graficos.pack(side="left", padx=10)
        
        self.var_ia = tk.BooleanVar(value=False)
        chk_ia = ttk.Checkbutton(main_frame, text="Incluir Análisis Asistido por IA (Llama-3)", variable=self.var_ia)
        chk_ia.grid(row=9, column=0, columnspan=2, **grid_lbl)
        
        self.var_obs = tk.BooleanVar(value=False)
        chk_obs = ttk.Checkbutton(main_frame, text="Incluir Notas u Observaciones Personalizadas", variable=self.var_obs, command=self.toggle_obs)
        chk_obs.grid(row=10, column=0, columnspan=2, **grid_lbl)
        
        ttk.Label(main_frame, text="Texto Observaciones:").grid(row=11, column=0, **grid_lbl)
        self.txt_obs = ttk.Entry(main_frame, state="disabled")
        self.txt_obs.grid(row=11, column=1, **grid_inp)
        
        self.btn_generar = ttk.Button(main_frame, text="GENERAR REPORTES", command=self.ejecutar_hilo)
        self.btn_generar.grid(row=12, column=0, columnspan=2, pady=12, sticky="ew")
        
        self.cb_modo.bind("<<ComboboxSelected>>", self.toggle_modo6)
        self.toggle_modo6()

    def cargar_excel_local(self):
        ruta = filedialog.askopenfilename(title="Seleccionar archivo Excel local", filetypes=[("Archivos Excel", "*.xlsx *.xls")])
        if ruta:
            try:
                shutil.copy(ruta, RUTA_EXCEL_LOCAL)
                self.root.after(0, lambda: messagebox.showinfo("Éxito", f"Excel local cargado correctamente.\n\nYa puedes cargar defectos y generar reportes precisos con tu versión actual."))
            except Exception as ex:
                self.root.after(0, lambda: messagebox.showerror("Error", f"No se pudo copiar el archivo: {ex}"))

    def toggle_modo6(self, event=None):
        if self.cb_modo.get().startswith("6"):
            self.lbl_modo6_text.grid(row=6, column=0, padx=10, pady=5, sticky="w")
            self.frame_list.grid(row=6, column=1, padx=10, pady=5, sticky="ew")
        else:
            self.lbl_modo6_text.grid_remove()
            self.frame_list.grid_remove()
            
    def toggle_obs(self):
        if self.var_obs.get():
            self.txt_obs.config(state="normal")
        else:
            self.txt_obs.delete(0, tk.END)
            self.txt_obs.config(state="disabled")
            
    def _on_map_tema(self, event):
        w = event.widget
        try:
            if w == w.winfo_toplevel():
                aplicar_tema_ventana(w)
        except Exception:
            pass
        
    def centrar_popup(self, popup, ancho, alto):
        self.root.update_idletasks()
        x_root, y_root = self.root.winfo_x(), self.root.winfo_y()
        w_root, h_root = self.root.winfo_width(), self.root.winfo_height()
        x = x_root + (w_root // 2) - (ancho // 2)
        y = y_root + (h_root // 2) - (alto // 2)
        popup.geometry(f"{ancho}x{alto}+{max(x, 0)}+{max(y, 0)}")
        
    def abrir_popup_graficos(self):
        opciones = [
            "Gráfico Principal (Según Modo)", "Gráfico Evolutivo (Tiempo)", "Diagrama de Pareto (80/20)",
            "Mapa de Calor (Área vs Defecto)", "Áreas Apiladas (Temporal)", "Distribución por Familia (Top 10)",
            "Rendimiento por Turno", "Ranking por Responsable de Turno", "Gráfico de Control (Tendencia)",
            "Comparativa entre Familias (Tasa %)", "Cruce Área x Turno (Mapa de Calor)"
        ]
        popup = tk.Toplevel(self.root)
        popup.title("Gráficos")
        bg_color = "#121212" if self.es_oscuro else "#f8fafc"
        popup.configure(bg=bg_color, bd=0, highlightthickness=0)
        self.centrar_popup(popup, 400, 500)
        lbl = ttk.Label(popup, text="Seleccione gráficos a incluir:", font=("Segoe UI", 10, "bold"), background=bg_color)
        lbl.pack(pady=10)
        self.vars_graficos = {}
        def update_graf():
            self.graficos_seleccionados = [g for g, v in self.vars_graficos.items() if v.get()]
            self.lbl_graficos.config(text=f"{len(self.graficos_seleccionados)} seleccionados")
        for g in opciones:
            var = tk.BooleanVar(value=(g in self.graficos_seleccionados))
            self.vars_graficos[g] = var
            ttk.Checkbutton(popup, text=g, variable=var, command=update_graf).pack(anchor="w", padx=20, pady=5)
            
    def hilo_sincronizar(self):
        self.btn_sync.config(state="disabled")
        threading.Thread(target=self.sincronizar_db, daemon=True).start()
        
    def sincronizar_db(self):
        try:
            url_sharepoint = "" #aca deberia ir la url que se consigue al compartir desde excel, se debe cambiar lo ultimo por ?download=1 para que descargue el archivo directamente
            response = requests.get(url_sharepoint, headers={'User-Agent': 'Mozilla/5.0'}, timeout=60)
            if response.status_code == 200:
                with open(RUTA_EXCEL_LOCAL, 'wb') as f:
                    f.write(response.content)
                self.root.after(0, lambda: messagebox.showinfo("Éxito", "Sincronización completada. Excel descargado de SharePoint correctamente."))
            else:
                self.root.after(0, lambda: messagebox.showerror("Error", "No se pudo conectar a SharePoint."))
        except Exception as ex:
            self.root.after(0, lambda: messagebox.showerror("Error", f"Fallo al descargar: {ex}"))
        finally:
            self.root.after(0, lambda: self.btn_sync.config(state="normal"))
            
    def hilo_cargar_defectos(self):
        if hasattr(self, 'defectos_disponibles') and self.defectos_disponibles:
            self.abrir_popup_seleccion(self.defectos_disponibles)
        else:
            self.btn_cargar.config(state="disabled")
            threading.Thread(target=self.cargar_defectos, daemon=True).start()
            
    def cargar_defectos(self):
        try:
            if not os.path.exists(RUTA_EXCEL_LOCAL):
                self.root.after(0, lambda: messagebox.showwarning("Aviso", "No hay archivo Excel local. Ejecuta la descarga o selecciona un Excel de tu PC primero."))
                return
            try:
                xls = pd.ExcelFile(RUTA_EXCEL_LOCAL)
                df_final = pd.DataFrame()
                hoja_pnc = next((s for s in xls.sheet_names if 'PNC' in s.upper()), None)
                if hoja_pnc:
                    header_idx = 0
                    df_temp = pd.read_excel(xls, sheet_name=hoja_pnc, header=None, nrows=15)
                    for i in range(len(df_temp)):
                        if 'PTA' in df_temp.iloc[i].astype(str).str.strip().str.upper().tolist():
                            header_idx = i
                            break
                    df_pnc = pd.read_excel(xls, sheet_name=hoja_pnc, header=header_idx)
                    df_pnc = limpiar_dataframe(df_pnc)
                    df_final = pd.concat([df_final, df_pnc], ignore_index=True)
                hoja_mp = next((s for s in xls.sheet_names if 'MP' in s.upper() and 'DATA' in s.upper()), None)
                if not hoja_mp:
                    hoja_mp = next((s for s in xls.sheet_names if 'RECHAZO MP' in s.upper()), None)
                if hoja_mp:
                    header_idx = 0
                    df_temp = pd.read_excel(xls, sheet_name=hoja_mp, header=None, nrows=15)
                    for i in range(len(df_temp)):
                        if 'PTA' in df_temp.iloc[i].astype(str).str.strip().str.upper().tolist():
                            header_idx = i
                            break
                    df_mp = pd.read_excel(xls, sheet_name=hoja_mp, header=header_idx)
                    df_mp = limpiar_dataframe(df_mp, tipo_incidencia='RECHAZO MP')
                    df_final = pd.concat([df_final, df_mp], ignore_index=True)
                if not df_final.empty:
                    col_def = 'DEFECTO' if 'DEFECTO' in df_final.columns else (df_final.columns[0] if not df_final.empty else 'DEFECTO')
                    if col_def in df_final.columns:
                        defectos = sorted([str(d).strip() for d in df_final[col_def].unique() if str(d).strip() and str(d).strip().lower() != 'nan'])
                        self.root.after(0, self.abrir_popup_seleccion, defectos)
                    else:
                        self.root.after(0, lambda: messagebox.showwarning("Aviso", "No se encontraron defectos procesables en el archivo."))
                else:
                    self.root.after(0, lambda: messagebox.showwarning("Aviso", "No se encontró la hoja 'PNC' ni 'MP' en el archivo Excel."))
            except Exception as ex:
                self.root.after(0, lambda: messagebox.showwarning("Aviso", f"Error procesando el Excel: {ex}"))
        except Exception as ex:
            registrar_error("Error cargar defectos Excel", ex)
        finally:
            self.root.after(0, lambda: self.btn_cargar.config(state="normal"))
            
    def abrir_popup_seleccion(self, defectos):
        self.defectos_disponibles = defectos
        popup = tk.Toplevel(self.root)
        popup.title("Selección")
        bg_color = "#121212" if self.es_oscuro else "#f8fafc"
        popup.configure(bg=bg_color, bd=0, highlightthickness=0)
        self.centrar_popup(popup, 450, 550)
        lbl = ttk.Label(popup, text="Marque las opciones:", font=("Segoe UI", 10, "bold"), background=bg_color)
        lbl.pack(pady=10)
        container = tk.Frame(popup, bg=bg_color, bd=0, highlightthickness=0)
        container.pack(fill="both", expand=True, padx=10, pady=5)
        canvas = tk.Canvas(container, bg=bg_color, highlightthickness=0, bd=0)
        scroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        frame_chk = tk.Frame(canvas, bg=bg_color, bd=0, highlightthickness=0)
        frame_chk.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frame_chk, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.vars_defectos = {}
        def update_def():
            self.defectos_seleccionados = [d for d, v in self.vars_defectos.items() if v.get()]
            self.lbl_seleccionados.config(text=f"{len(self.defectos_seleccionados)} seleccionados")
        for d in defectos:
            var = tk.BooleanVar(value=(d in self.defectos_seleccionados))
            self.vars_defectos[d] = var
            chk = ttk.Checkbutton(frame_chk, text=d, variable=var, command=update_def)
            chk.pack(anchor="w", padx=5, pady=2)
            
    def ejecutar_hilo(self):
        if any(not str(c.get()).strip() for c in [self.cb_modo, self.cb_tipo, self.cb_turno, self.cb_formato]):
            messagebox.showwarning("Selección Requerida", "Por favor seleccione todas las opciones requeridas.")
            return
        if not os.path.exists(RUTA_EXCEL_LOCAL):
            messagebox.showwarning("Archivo Faltante", "No se encontró el Excel local. Usa el botón 'Cargar Excel Local' o sincroniza desde SharePoint primero.")
            return
        self.btn_generar.config(state="disabled")
        threading.Thread(target=self.procesar_reporte, daemon=True).start()
        
    def solicitar_produccion_popup(self):
        res = [None]
        def ask():
            res[0] = simpledialog.askfloat("Dato Requerido", "No se encontró el valor de la Producción Real.\nPor favor ingrese la cantidad total:", parent=self.root)
        self.root.after(0, ask)
        while res[0] is None:
            time.sleep(0.1)
        return res[0]
        
    def obtener_api_key_groq(self):
        key = os.environ.get("GROQ_API_KEY")
        if key: return key
        _sin_respuesta = object()
        res = [_sin_respuesta]
        def ask():
            res[0] = simpledialog.askstring("Configuración requerida", "Ingresa la API Key de Groq para el Análisis IA.\nSolo se pedirá una vez en este equipo.", parent=self.root, show="*")
        self.root.after(0, ask)
        while res[0] is _sin_respuesta:
            time.sleep(0.1)
        key = res[0]
        if key:
            try:
                subprocess.run(["setx", "GROQ_API_KEY", key], shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as ex:
                registrar_error("No se pudo guardar la API Key", ex)
            os.environ["GROQ_API_KEY"] = key
        return key
        
    def procesar_reporte(self):
        try:
            autor_nombre = self.txt_autor.get().strip()
            if autor_nombre:
                try:
                    with open(obtener_ruta_config_autor(), 'w', encoding='utf-8') as f:
                        f.write(autor_nombre)
                except Exception:
                    pass
            modo_rep = self.cb_modo.get().split(" - ")[0].strip()
            tipo_sel = self.cb_tipo.get().split(" - ")[0].strip()
            turno_sel = self.cb_turno.get().split(" - ")[0].strip()
            f_in_str = self.txt_f_in.get().strip()
            f_out_str = self.txt_f_out.get().strip()
            seleccionados = getattr(self, 'defectos_seleccionados', [])
            
            if modo_rep == '6' and not seleccionados:
                self.btn_generar.config(state="normal")
                return
                
            if not f_in_str:
                fin, fout = pd.to_datetime('2000-01-01'), pd.to_datetime('2100-01-01')
                txt_rango, file_rango = "Histórico Completo", "Historico"
            else:
                anio_actual = datetime.now().year
                f_in_str = f_in_str.replace('/', '-')
                if len(f_in_str) <= 5:
                    f_in_str += f"-{anio_actual}"
                fin = pd.to_datetime(f_in_str, format='%d-%m-%Y')
                if not f_out_str:
                    fout = fin
                    txt_rango, file_rango = fin.strftime('%d-%m-%Y'), fin.strftime('%d-%m-%Y')
                else:
                    f_out_str = f_out_str.replace('/', '-')
                    if len(f_out_str) <= 5:
                        f_out_str += f"-{anio_actual}"
                    fout = pd.to_datetime(f_out_str, format='%d-%m-%Y')
                    txt_rango, file_rango = f"Del {fin.strftime('%d-%m-%Y')} al {fout.strftime('%d-%m-%Y')}", f"{fin.strftime('%d-%m-%Y')}_al_{fout.strftime('%d-%m-%Y')}"
                    
            if turno_sel == '2':
                filtro_turno, txt_turno, file_turno = 'Turno A', 'Turno A (Día)', 'TA'
            elif turno_sel == '3':
                filtro_turno, txt_turno, file_turno = 'Turno B', 'Turno B (Noche)', 'TB'
            else:
                filtro_turno, txt_turno, file_turno = None, '', '24H'
            
            if tipo_sel == '1': tipo_str, titulo_reporte, color_hex, color_hex_end = 'Rechazo', 'RECHAZOS', '#ef4444', '#7f1d1d'
            elif tipo_sel == '2': tipo_str, titulo_reporte, color_hex, color_hex_end = 'Reparacion', 'REPARACIONES', '#008778', '#045c4f'
            elif tipo_sel == '3': tipo_str, titulo_reporte, color_hex, color_hex_end = 'Rechazo_MP', 'RECHAZOS MP', '#d97706', '#78350f'
            elif tipo_sel == '4': tipo_str, titulo_reporte, color_hex, color_hex_end = 'Rech_y_Rep', 'RECHAZOS Y REPARACIONES', '#ef4444', '#008778'
            elif tipo_sel == '5': tipo_str, titulo_reporte, color_hex, color_hex_end = 'Consolidado', 'CONSOLIDADO', '#ef4444', '#008778'
            else: tipo_str, titulo_reporte, color_hex, color_hex_end = 'Reporte', 'REPORTE', '#008778', '#045c4f'
            
            formato_salida = '1' if self.cb_formato.get() == "PDF" else ('2' if self.cb_formato.get() == "HTML" else '3')
            usar_ia, usar_obs = self.var_ia.get(), self.var_obs.get()
            observaciones_texto = self.txt_obs.get().strip() if usar_obs else ""
            ruta_escritorio = obtener_ruta_escritorio()
            
            df_pnc_raw = pd.DataFrame()
            total_meta_real, total_entregado_real = 0.0, 0.0
            
            try:
                xls = pd.ExcelFile(RUTA_EXCEL_LOCAL)
                df_final = pd.DataFrame()
                
                hoja_pnc = next((s for s in xls.sheet_names if 'PNC' in s.upper()), None)
                if hoja_pnc:
                    header_idx = 0
                    df_temp = pd.read_excel(xls, sheet_name=hoja_pnc, header=None, nrows=15)
                    for i in range(len(df_temp)):
                        if 'PTA' in df_temp.iloc[i].astype(str).str.strip().str.upper().tolist():
                            header_idx = i
                            break
                    df_raw = pd.read_excel(xls, sheet_name=hoja_pnc, header=header_idx)
                    df_pnc_limpio = limpiar_dataframe(df_raw)
                    df_final = pd.concat([df_final, df_pnc_limpio], ignore_index=True)
                    
                hoja_mp = next((s for s in xls.sheet_names if 'MP' in s.upper() and 'DATA' in s.upper()), None)
                if not hoja_mp:
                    hoja_mp = next((s for s in xls.sheet_names if 'RECHAZO MP' in s.upper()), None)
                    
                if hoja_mp:
                    header_idx = 0
                    df_temp = pd.read_excel(xls, sheet_name=hoja_mp, header=None, nrows=15)
                    for i in range(len(df_temp)):
                        if 'PTA' in df_temp.iloc[i].astype(str).str.strip().str.upper().tolist():
                            header_idx = i
                            break
                    df_mp_raw = pd.read_excel(xls, sheet_name=hoja_mp, header=header_idx)
                    df_mp_limpio = limpiar_dataframe(df_mp_raw, tipo_incidencia='RECHAZO MP')
                    df_final = pd.concat([df_final, df_mp_limpio], ignore_index=True)
                    
                df_pnc_raw = df_final
                
                if 'CUMPLIMIENTO MENSUAL' in xls.sheet_names:
                    df_cump_raw = pd.read_excel(xls, sheet_name='CUMPLIMIENTO MENSUAL', header=None)
                    sub_mes = df_cump_raw.iloc[:, 12:15].dropna(how='all').copy()
                    sub_mes.columns = ['CANAL', 'META', 'REAL']
                    for _, row in sub_mes.iterrows():
                        v_canal = str(row['CANAL']).strip().lower() if pd.notna(row['CANAL']) else ''
                        if v_canal == 'total':
                            try:
                                total_meta_real = float(row['META'])
                                total_entregado_real = float(row['REAL'])
                                break
                            except Exception:
                                pass
            except Exception as ex:
                registrar_error("Error procesando excel local", ex)
                
            if not df_pnc_raw.empty and 'CANTIDAD' in df_pnc_raw.columns:
                df_pnc_raw['CANTIDAD_FISICA'] = df_pnc_raw['CANTIDAD']
            
            if total_entregado_real <= 0 and modo_rep == '2':
                val = self.solicitar_produccion_popup()
                if val and val > 0:
                    total_entregado_real = float(val)
                
            if tipo_sel == '1': df_master = df_pnc_raw[df_pnc_raw['TIPO_INCIDENCIA'] == 'RECHAZO'].copy() if not df_pnc_raw.empty else pd.DataFrame()
            elif tipo_sel == '2': df_master = df_pnc_raw[df_pnc_raw['TIPO_INCIDENCIA'] == 'REPARACION'].copy() if not df_pnc_raw.empty else pd.DataFrame()
            elif tipo_sel == '3': df_master = df_pnc_raw[df_pnc_raw['TIPO_INCIDENCIA'] == 'RECHAZO MP'].copy() if not df_pnc_raw.empty else pd.DataFrame()
            elif tipo_sel == '4': df_master = df_pnc_raw[df_pnc_raw['TIPO_INCIDENCIA'].isin(['RECHAZO', 'REPARACION'])].copy() if not df_pnc_raw.empty else pd.DataFrame()
            elif tipo_sel == '5': df_master = df_pnc_raw.copy()
            else: df_master = df_pnc_raw.copy()
            
            df_historico_chart = pd.DataFrame()
            if not df_master.empty and 'FECHA REVISION' in df_master.columns:
                df_master['FECHA REVISION'] = pd.to_datetime(df_master['FECHA REVISION'], errors='coerce')
                df_f = df_master[(df_master['FECHA REVISION'] >= fin) & (df_master['FECHA REVISION'] <= fout)].copy()
                df_historico_chart = df_master[df_master['FECHA REVISION'] <= fout].copy()
            else: 
                df_f = df_master.copy()
                df_historico_chart = df_master.copy()
                
            if filtro_turno and 'TURNO' in df_f.columns: 
                df_f = df_f[df_f['TURNO'] == filtro_turno]
                if 'TURNO' in df_historico_chart.columns:
                    df_historico_chart = df_historico_chart[df_historico_chart['TURNO'] == filtro_turno]
            
            mapa_modos = {'1': 'EST', '2': 'KPI', '3': 'A6M', '4': 'EDF', '5': 'FAM', '6': 'CED'}
            out_name = f"Reporte_{mapa_modos.get(modo_rep, 'REP')}_{tipo_str}_{file_rango}_{file_turno}.pdf"
            
            col_def = 'DEFECTO' if 'DEFECTO' in df_f.columns else (df_f.columns[0] if not df_f.empty else 'DEFECTO')
            col_causa = 'POSIBLE CAUSA RAIZ' if 'POSIBLE CAUSA RAIZ' in df_f.columns else ('CAUSA' if 'CAUSA' in df_f.columns else 'DEFECTO')
            col_fam = 'FAMILIA' if 'FAMILIA' in df_f.columns else col_def
            
            def get_top_str(df_sub, col):
                try: return str(df_sub.groupby(col)['CANTIDAD_FISICA'].sum().idxmax())
                except Exception: return "N/A"
                
            def get_top_pair(df_sub, col_a, col_b):
                top_a = get_top_str(df_sub, col_a)
                if top_a == "N/A": return top_a, "N/A"
                top_b = get_top_str(df_sub[df_sub[col_a] == top_a], col_b)
                return top_a, top_b
                
            top_def_hist = get_top_str(df_f, col_def)
            if modo_rep == '4' and top_def_hist != "N/A":
                df_base_tabla = df_f[df_f[col_def] == top_def_hist].copy()
            elif modo_rep == '6' and seleccionados:
                df_base_tabla = df_f[df_f[col_def].isin(seleccionados)].copy()
            else:
                df_base_tabla = df_f.copy()
            
            qty_rechazos = df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'RECHAZO']['CANTIDAD_FISICA'].sum() if not df_base_tabla.empty and 'TIPO_INCIDENCIA' in df_base_tabla.columns else 0
            qty_reparaciones = df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'REPARACION']['CANTIDAD_FISICA'].sum() if not df_base_tabla.empty and 'TIPO_INCIDENCIA' in df_base_tabla.columns else 0
            qty_mp = df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'RECHAZO MP']['CANTIDAD_FISICA'].sum() if not df_base_tabla.empty and 'TIPO_INCIDENCIA' in df_base_tabla.columns else 0
            total_unid = qty_rechazos + qty_reparaciones + qty_mp
            
            top_def, top_causa = get_top_pair(df_base_tabla, col_def, col_causa)
            top_def_rech, top_causa_rech = get_top_pair(df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'RECHAZO'], col_def, col_causa)
            top_def_rep, top_causa_rep = get_top_pair(df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'REPARACION'], col_def, col_causa)
            top_def_mp, top_causa_mp = get_top_pair(df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'] == 'RECHAZO MP'], col_def, col_causa)
            
            try:
                total_tipos = len(df_base_tabla[col_def].unique()) if modo_rep != '5' else len(df_base_tabla[col_fam].unique())
            except Exception:
                total_tipos = 0
            
            mostrar_rechazos = tipo_sel in ['1', '3', '4', '5']
            mostrar_reparaciones = tipo_sel in ['2', '4', '5']
            kpi_cols = 2 + (1 if mostrar_rechazos else 0) + (1 if mostrar_reparaciones else 0)
            
            cumplimiento_pct = (total_entregado_real / total_meta_real * 100) if total_meta_real > 0 else 0.0
            cumplimiento_str = f"{cumplimiento_pct:.1f}%"
            tasa_incidencias_calc = ((qty_rechazos + qty_reparaciones + qty_mp) / total_entregado_real * 100) if total_entregado_real > 0 else 0.0
            tasa_kpi_str = f"{tasa_incidencias_calc:.2f}%"
            
            color_borde_cump_hex = "#10b981" if cumplimiento_pct >= 100 else ("#C4D70F" if cumplimiento_pct >= 90 else "#ef4444")
            bg_cump = "#f0fdf4" if cumplimiento_pct >= 100 else ("#f7fee7" if cumplimiento_pct >= 90 else "#fef2f2")
            
            if fin != pd.to_datetime('2000-01-01') and fin <= fout:
                d_diff = (fout - fin).days + 1
                if d_diff < 7:
                    fechas_chart = pd.date_range(end=fout, periods=7)
                else:
                    fechas_chart = pd.date_range(start=fin, end=fout)
            else:
                fechas_chart = pd.date_range(end=datetime.now().normalize(), periods=7)
                
            lbls_chart = [d.strftime('%d-%m') for d in fechas_chart]
            img_chart1, img_chart2 = "", ""
            titulo_grafico1, titulo_grafico2, top_m = "", "", "N/A"
            chart2_ds = []
            
            if modo_rep in ['1', '2']:
                if not df_base_tabla.empty and 'TIPO_INCIDENCIA' in df_base_tabla.columns:
                    gf = df_base_tabla.groupby([col_def, 'TIPO_INCIDENCIA'])['CANTIDAD_FISICA'].sum().unstack(fill_value=0).reset_index()
                    gf['TOTAL'] = gf.sum(axis=1, numeric_only=True)
                    top10 = gf.sort_values(by='TOTAL', ascending=False).head(10)
                    labels = top10[col_def].astype(str).tolist()
                    ds1 = top10.get('RECHAZO', pd.Series([0] * len(labels))).tolist()
                    ds2 = top10.get('REPARACION', pd.Series([0] * len(labels))).tolist()
                    ds3 = top10.get('RECHAZO MP', pd.Series([0] * len(labels))).tolist()
                    chart1_ds = []
                    if sum(ds1) > 0 and mostrar_rechazos: chart1_ds.append({'label': 'Rechazo', 'data': ds1, 'color': '#ef4444'})
                    if sum(ds2) > 0 and mostrar_reparaciones: chart1_ds.append({'label': 'Reparación', 'data': ds2, 'color': '#008778'})
                    if sum(ds3) > 0 and mostrar_rechazos: chart1_ds.append({'label': 'Rechazo MP', 'data': ds3, 'color': '#d97706'})
                    if chart1_ds:
                        titulo_grafico1 = "Top 10 Incidencias"
                        img_chart1 = plot_bar(labels, chart1_ds, stacked=True)
                        
                ds_rech_chart, ds_rep_chart, ds_mp_chart = [0] * len(fechas_chart), [0] * len(fechas_chart), [0] * len(fechas_chart)
                for i, dt in enumerate(fechas_chart):
                    df_day = df_historico_chart[df_historico_chart['FECHA REVISION'].dt.normalize() == dt.normalize()]
                    if not df_day.empty and 'TIPO_INCIDENCIA' in df_day.columns:
                        ds_rech_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'RECHAZO']['CANTIDAD_FISICA'].sum()
                        ds_rep_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'REPARACION']['CANTIDAD_FISICA'].sum()
                        ds_mp_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'RECHAZO MP']['CANTIDAD_FISICA'].sum()
                if sum(ds_rech_chart) > 0 and mostrar_rechazos: chart2_ds.append({'label': 'Rechazo', 'data': ds_rech_chart, 'color': '#ef4444'})
                if sum(ds_rep_chart) > 0 and mostrar_reparaciones: chart2_ds.append({'label': 'Reparación', 'data': ds_rep_chart, 'color': '#008778'})
                if sum(ds_mp_chart) > 0 and mostrar_rechazos: chart2_ds.append({'label': 'Rechazo MP', 'data': ds_mp_chart, 'color': '#d97706'})
                if chart2_ds:
                    titulo_grafico2 = f"Evolución Tiempo ({len(fechas_chart)} Días)"
                    img_chart2 = plot_line(lbls_chart, chart2_ds)
                    
            elif modo_rep == '3':
                if not df_base_tabla.empty:
                    df_base_tabla['6M'] = df_base_tabla[col_causa].apply(clasificar_6m)
                    res_6m = df_base_tabla.groupby('6M')['CANTIDAD_FISICA'].sum().sort_values(ascending=False)
                    if not res_6m.empty:
                        top_m = str(res_6m.idxmax())
                        titulo_grafico1 = "Distribución Causa Raíz (6M)"
                        img_chart1 = plot_pie(res_6m.index.tolist(), res_6m.values.tolist())
                    gf = df_base_tabla.groupby([col_def])['CANTIDAD_FISICA'].sum().reset_index()
                    top10 = gf.sort_values(by='CANTIDAD_FISICA', ascending=False).head(10)
                    titulo_grafico2 = "Top 10 Incidencias"
                    img_chart2 = plot_bar(top10[col_def].astype(str).tolist(), [{'label': 'Volumen', 'data': top10['CANTIDAD_FISICA'].tolist(), 'color': color_hex}], stacked=False)
                    
            elif modo_rep == '4':
                if top_def != "N/A" and not df_base_tabla.empty:
                    ds_def_chart = [0] * len(fechas_chart)
                    for i, dt in enumerate(fechas_chart):
                        df_day = df_historico_chart[(df_historico_chart['FECHA REVISION'].dt.normalize() == dt.normalize()) & (df_historico_chart[col_def] == top_def_hist)]
                        if not df_day.empty:
                            ds_def_chart[i] = df_day['CANTIDAD_FISICA'].sum()
                    chart2_ds = [{'label': top_def, 'data': ds_def_chart, 'color': color_hex}]
                    titulo_grafico1 = f"Evolución: {top_def} ({len(fechas_chart)} Días)"
                    img_chart1 = plot_line(lbls_chart, chart2_ds)
                    gf = df_base_tabla.groupby(col_causa)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(5)
                    titulo_grafico2 = f"Top 5 Causas Asociadas a: {top_def}"
                    img_chart2 = plot_bar(gf[col_causa].astype(str).tolist(), [{'label': 'Causas', 'data': gf['CANTIDAD_FISICA'].tolist(), 'color': color_hex}], stacked=False)
                    
            elif modo_rep == '5':
                if not df_base_tabla.empty and col_fam in df_base_tabla.columns:
                    gf = df_base_tabla.groupby([col_fam, 'TIPO_INCIDENCIA'])['CANTIDAD_FISICA'].sum().unstack(fill_value=0).reset_index()
                    gf['TOTAL'] = gf.sum(axis=1, numeric_only=True)
                    top10 = gf.sort_values(by='TOTAL', ascending=False).head(10)
                    labels = top10[col_fam].astype(str).tolist()
                    ds1 = top10.get('RECHAZO', pd.Series([0] * len(labels))).tolist()
                    ds2 = top10.get('REPARACION', pd.Series([0] * len(labels))).tolist()
                    ds3 = top10.get('RECHAZO MP', pd.Series([0] * len(labels))).tolist()
                    chart1_ds = []
                    if sum(ds1) > 0 and mostrar_rechazos: chart1_ds.append({'label': 'Rechazo', 'data': ds1, 'color': '#ef4444'})
                    if sum(ds2) > 0 and mostrar_reparaciones: chart1_ds.append({'label': 'Reparación', 'data': ds2, 'color': '#008778'})
                    if sum(ds3) > 0 and mostrar_rechazos: chart1_ds.append({'label': 'Rechazo MP', 'data': ds3, 'color': '#d97706'})
                    if chart1_ds:
                        titulo_grafico1 = "Top Incidencias por Familia"
                        img_chart1 = plot_bar(labels, chart1_ds, stacked=True)
                        
                ds_rech_chart, ds_rep_chart, ds_mp_chart = [0] * len(fechas_chart), [0] * len(fechas_chart), [0] * len(fechas_chart)
                for i, dt in enumerate(fechas_chart):
                    df_day = df_historico_chart[df_historico_chart['FECHA REVISION'].dt.normalize() == dt.normalize()]
                    if not df_day.empty and 'TIPO_INCIDENCIA' in df_day.columns:
                        ds_rech_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'RECHAZO']['CANTIDAD_FISICA'].sum()
                        ds_rep_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'REPARACION']['CANTIDAD_FISICA'].sum()
                        ds_mp_chart[i] = df_day[df_day['TIPO_INCIDENCIA'] == 'RECHAZO MP']['CANTIDAD_FISICA'].sum()
                if sum(ds_rech_chart) > 0 and mostrar_rechazos: chart2_ds.append({'label': 'Rechazo', 'data': ds_rech_chart, 'color': '#ef4444'})
                if sum(ds_rep_chart) > 0 and mostrar_reparaciones: chart2_ds.append({'label': 'Reparación', 'data': ds_rep_chart, 'color': '#008778'})
                if sum(ds_mp_chart) > 0 and mostrar_rechazos: chart2_ds.append({'label': 'Rechazo MP', 'data': ds_mp_chart, 'color': '#d97706'})
                if chart2_ds:
                    titulo_grafico2 = f"Evolución Tiempo ({len(fechas_chart)} Días)"
                    img_chart2 = plot_line(lbls_chart, chart2_ds)
                    
            elif modo_rep == '6':
                top_m = str(len(seleccionados))
                colores = ['#ef4444', '#008778', '#F5911E', '#3b82f6', '#8b5cf6', '#14b8a6', '#f59e0b', '#10b981', '#6366f1']
                if not df_base_tabla.empty:
                    for idx, defecto in enumerate(seleccionados):
                        df_d = df_historico_chart[df_historico_chart[col_def] == defecto]
                        ds_data = [0] * len(fechas_chart)
                        for i, dt in enumerate(fechas_chart):
                            df_day = df_d[df_d['FECHA REVISION'].dt.normalize() == dt.normalize()]
                            if not df_day.empty:
                                ds_data[i] = df_day['CANTIDAD_FISICA'].sum()
                        if sum(ds_data) > 0:
                            chart2_ds.append({'label': str(defecto)[:15], 'data': ds_data, 'color': colores[idx % len(colores)]})
                    if chart2_ds:
                        titulo_grafico1 = "Evolución Comparativa Seleccionados"
                        img_chart1 = plot_line(lbls_chart, chart2_ds)
                    gf = df_base_tabla.groupby(col_def)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(10)
                    if not gf.empty:
                        titulo_grafico2 = "Volumen Total por Defecto Seleccionado"
                        img_chart2 = plot_bar(gf[col_def].astype(str).tolist(), [{'label': 'Volumen', 'data': gf['CANTIDAD_FISICA'].tolist(), 'color': color_hex}], stacked=False)
                        
            graficos_dinamicos = []
            if "Gráfico Principal (Según Modo)" in getattr(self, 'graficos_seleccionados', []) and img_chart1:
                graficos_dinamicos.append({'titulo': titulo_grafico1, 'b64': img_chart1})
            if "Gráfico Evolutivo (Tiempo)" in getattr(self, 'graficos_seleccionados', []) and img_chart2:
                graficos_dinamicos.append({'titulo': titulo_grafico2, 'b64': img_chart2})
            if "Diagrama de Pareto (80/20)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty:
                col_obj = col_fam if modo_rep == '5' else col_def
                gf_pareto = df_base_tabla.groupby(col_obj)['CANTIDAD_FISICA'].sum().reset_index()
                gf_pareto = gf_pareto[gf_pareto['CANTIDAD_FISICA'] > 0]
                if not gf_pareto.empty:
                    img_pareto = plot_pareto(gf_pareto[col_obj].astype(str).tolist(), gf_pareto['CANTIDAD_FISICA'].tolist())
                    graficos_dinamicos.append({'titulo': f'Diagrama de Pareto (80/20) - {col_obj}', 'b64': img_pareto})
            if "Mapa de Calor (Área vs Defecto)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and 'AREA' in df_base_tabla.columns:
                col_obj = col_fam if modo_rep == '5' else col_def
                img_heat = plot_heatmap(df_base_tabla, 'AREA', col_obj)
                if img_heat:
                    graficos_dinamicos.append({'titulo': f'Mapa de Calor - Área vs {col_obj}', 'b64': img_heat})
            if "Áreas Apiladas (Temporal)" in getattr(self, 'graficos_seleccionados', []) and modo_rep != '3' and chart2_ds:
                img_area = plot_stacked_area(lbls_chart, chart2_ds)
                if img_area:
                    graficos_dinamicos.append({'titulo': 'Evolución Temporal Apilada', 'b64': img_area})
            if "Distribución por Familia (Top 10)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and col_fam in df_base_tabla.columns:
                gf_fam = df_base_tabla.groupby(col_fam)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(10)
                if not gf_fam.empty:
                    img_fam = plot_bar(gf_fam[col_fam].astype(str).tolist(), [{'label': 'Volumen', 'data': gf_fam['CANTIDAD_FISICA'].tolist(), 'color': color_hex}], stacked=False)
                    if img_fam:
                        graficos_dinamicos.append({'titulo': "Top 10 Familias Afectadas", 'b64': img_fam})
            if "Rendimiento por Turno" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and 'TURNO' in df_base_tabla.columns:
                gf_turno = df_base_tabla.groupby(['TURNO', 'TIPO_INCIDENCIA'])['CANTIDAD_FISICA'].sum().unstack(fill_value=0)
                if not gf_turno.empty:
                    turnos_lbl = gf_turno.index.astype(str).tolist()
                    ds_turno = []
                    if 'RECHAZO' in gf_turno.columns and mostrar_rechazos:
                        ds_turno.append({'label': 'Rechazo', 'data': gf_turno['RECHAZO'].tolist(), 'color': '#ef4444'})
                    if 'REPARACION' in gf_turno.columns and mostrar_reparaciones:
                        ds_turno.append({'label': 'Reparación', 'data': gf_turno['REPARACION'].tolist(), 'color': '#008778'})
                    if 'RECHAZO MP' in gf_turno.columns and mostrar_rechazos:
                        ds_turno.append({'label': 'Rechazo MP', 'data': gf_turno['RECHAZO MP'].tolist(), 'color': '#d97706'})
                    if ds_turno:
                        img_turno = plot_bar(turnos_lbl, ds_turno, stacked=False)
                        if img_turno:
                            graficos_dinamicos.append({'titulo': 'Rendimiento por Turno', 'b64': img_turno})
            if "Ranking por Responsable de Turno" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty:
                col_resp = 'RESPONSABLE TURNO' if 'RESPONSABLE TURNO' in df_base_tabla.columns else ('RESPONSABLE' if 'RESPONSABLE' in df_base_tabla.columns else None)
                if col_resp:
                    gf_resp = df_base_tabla.groupby(col_resp)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(10)
                    if not gf_resp.empty:
                        img_resp = plot_hbar(gf_resp[col_resp].astype(str).tolist(), gf_resp['CANTIDAD_FISICA'].tolist())
                        if img_resp:
                            graficos_dinamicos.append({'titulo': 'Ranking por Responsable de Turno', 'b64': img_resp})
            if "Gráfico de Control (Tendencia)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and 'FECHA REVISION' in df_base_tabla.columns:
                ds_control = [0] * len(fechas_chart)
                for i, dt in enumerate(fechas_chart):
                    df_day = df_historico_chart[df_historico_chart['FECHA REVISION'].dt.normalize() == dt.normalize()]
                    ds_control[i] = df_day['CANTIDAD_FISICA'].sum() if not df_day.empty else 0
                if sum(ds_control) > 0:
                    img_control = plot_control(lbls_chart, ds_control)
                    if img_control:
                        graficos_dinamicos.append({'titulo': f'Gráfico de Control ({len(fechas_chart)} Días)', 'b64': img_control})
            if "Comparativa entre Familias (Tasa %)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and col_fam in df_base_tabla.columns:
                gf_fam_tot = df_base_tabla.groupby(col_fam)['CANTIDAD_FISICA'].sum().reset_index().rename(columns={'CANTIDAD_FISICA': 'TOTAL'})
                gf_fam_rech = df_base_tabla[df_base_tabla['TIPO_INCIDENCIA'].isin(['RECHAZO', 'RECHAZO MP'])].groupby(col_fam)['CANTIDAD_FISICA'].sum().reset_index().rename(columns={'CANTIDAD_FISICA': 'RECHAZO'})
                gf_tasa = gf_fam_tot.merge(gf_fam_rech, on=col_fam, how='left').fillna(0)
                gf_tasa = gf_tasa[gf_tasa['TOTAL'] > 0]
                gf_tasa['TASA'] = (gf_tasa['RECHAZO'] / gf_tasa['TOTAL'] * 100).round(1)
                gf_tasa = gf_tasa.sort_values(by='TASA', ascending=False).head(10)
                if not gf_tasa.empty:
                    img_tasa = plot_bar(gf_tasa[col_fam].astype(str).tolist(), [{'label': 'Tasa Rechazo %', 'data': gf_tasa['TASA'].tolist(), 'color': '#ef4444'}], stacked=False)
                    if img_tasa:
                        graficos_dinamicos.append({'titulo': 'Comparativa entre Familias (Tasa % Rechazo)', 'b64': img_tasa})
            if "Cruce Área x Turno (Mapa de Calor)" in getattr(self, 'graficos_seleccionados', []) and not df_base_tabla.empty and 'AREA' in df_base_tabla.columns and 'TURNO' in df_base_tabla.columns:
                img_heat_turno = plot_heatmap(df_base_tabla, 'AREA', 'TURNO')
                if img_heat_turno:
                    graficos_dinamicos.append({'titulo': 'Cruce Área x Turno', 'b64': img_heat_turno})
                    
            df_tabla = pd.DataFrame({
                'TIPO': df_base_tabla.get('TIPO_INCIDENCIA', ''),
                'FECHA': df_base_tabla['FECHA REVISION'].dt.strftime('%d-%m-%Y') if not df_base_tabla.empty and pd.api.types.is_datetime64_any_dtype(df_base_tabla.get('FECHA REVISION')) else df_base_tabla.get('FECHA REVISION', ''),
                'AREA': df_base_tabla.get('AREA', ''),
                'ORDEN': df_base_tabla.get('ORDEN', ''),
                'PTA': df_base_tabla.get('PTA', ''),
                'DEFECTO': df_base_tabla.get(col_fam if modo_rep == '5' else col_def, ''),
                'CAUSA': df_base_tabla.get(col_causa, ''),
                'CANT': df_base_tabla.apply(lambda r: str(int(r.get('CANTIDAD_FISICA', 0))), axis=1) if not df_base_tabla.empty else []
            })
            
            resumen_ia = ""
            if usar_ia and not df_base_tabla.empty:
                _total_inc_ia = int(qty_rechazos + qty_reparaciones + qty_mp)
                _col_causa_ia = col_causa if col_causa in df_base_tabla.columns else col_def
                _top_def_ia = df_base_tabla.groupby(col_def)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(3)
                _top_causa_ia = df_base_tabla.groupby(_col_causa_ia)['CANTIDAD_FISICA'].sum().reset_index().sort_values(by='CANTIDAD_FISICA', ascending=False).head(3)
                contexto_ia = {
                    'periodo': txt_rango,
                    'tasa_incidencias_actual_pct': round(tasa_incidencias_calc, 2),
                    'cumplimiento_meta_pct': round(cumplimiento_pct, 1),
                    'total_rechazos': int(qty_rechazos + qty_mp),
                    'total_reparaciones': int(qty_reparaciones),
                    'top_defectos_pct_del_total': [{'nombre': str(r[col_def]), 'unidades': int(r['CANTIDAD_FISICA']), 'pct': round(r['CANTIDAD_FISICA'] / _total_inc_ia * 100, 1) if _total_inc_ia > 0 else 0} for _, r in _top_def_ia.iterrows()],
                    'top_causas_pct_del_total': [{'nombre': str(r[_col_causa_ia]), 'unidades': int(r['CANTIDAD_FISICA']), 'pct': round(r['CANTIDAD_FISICA'] / _total_inc_ia * 100, 1) if _total_inc_ia > 0 else 0} for _, r in _top_causa_ia.iterrows()]
                }
                _api_key_groq = self.obtener_api_key_groq()
                resumen_ia = redactar_analisis_ia(str(contexto_ia), _api_key_groq)
                
            html_final = Template(HTML_TEMPLATE).render(
                modo=modo_rep, autor_nombre=autor_nombre, usar_ia=usar_ia, analisis_ia=resumen_ia, usar_obs=usar_obs, observaciones=observaciones_texto,
                total_unid=int(total_unid), total_tipos=int(total_tipos), top_def=top_def, top_causa=top_causa, top_m=top_m,
                top_def_rech=top_def_rech, top_causa_rech=top_causa_rech, top_def_rep=top_def_rep, top_causa_rep=top_causa_rep,
                top_def_mp=top_def_mp, top_causa_mp=top_causa_mp, total_rechazos=int(qty_rechazos + qty_mp), total_reparaciones=int(qty_reparaciones),
                periodo=txt_rango, subtitulo_turno=txt_turno, fecha_generacion=datetime.now().strftime('%d-%m-%Y %H:%M'),
                datos_tabla=df_tabla.to_dict(orient='records'), total_meta=int(total_meta_real), total_entregado=int(total_entregado_real),
                cumplimiento_str=cumplimiento_str, tasa_produccion_str=f"{tasa_incidencias_calc:.2f}%", tasa_kpi_str=tasa_kpi_str,
                tasa_rechazo_str=f"{((qty_rechazos + qty_mp) / total_entregado_real * 100):.2f}%" if total_entregado_real > 0 else "0.0%",
                tasa_reparacion_str=f"{(qty_reparaciones / total_entregado_real * 100):.2f}%" if total_entregado_real > 0 else "0.0%",
                color_borde_cump_hex=color_borde_cump_hex, bg_cump=bg_cump, graficos_dinamicos=graficos_dinamicos,
                titulo_reporte=titulo_reporte, color_hex=color_hex, color_hex_end=color_hex_end,
                mostrar_rechazos=mostrar_rechazos, mostrar_reparaciones=mostrar_reparaciones, kpi_cols=kpi_cols
            )
            
            exportar_excel_calidad(df_base_tabla, df_pnc_raw, ruta_escritorio, out_name.replace('.pdf', ''))
            temp_dir = tempfile.gettempdir()
            out_html = os.path.join(temp_dir, "Temp_Report_Calidad.html")
            
            with open(out_html, 'w', encoding='utf-8') as f:
                f.write(html_final)
                
            out_pdf = os.path.join(ruta_escritorio, out_name)
            
            if formato_salida in ['1', '3']:
                convertir_html_a_pdf(out_html, out_pdf)
                aplicar_metadata_pdf(out_pdf, autor_nombre)
            if formato_salida in ['2', '3']:
                shutil.copy(out_html, os.path.join(ruta_escritorio, out_name.replace('.pdf', '.html')))
                
            time.sleep(0.5)
            if os.path.exists(out_html):
                try: os.remove(out_html)
                except Exception: pass
            messagebox.showinfo("Completado", f"Informe generado exitosamente.\n\nNombre: {out_name}")
            
        except Exception as ex:
            messagebox.showerror("Error", f"Ocurrió un error: {ex}")
        finally:
            self.btn_generar.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    app = PanelControlMafor(root)
    root.mainloop()